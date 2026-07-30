"""
数据导出模块
支持导出为 CSV、JSON、Excel 格式，兼容大乐透和双色球
"""
import csv
import json
import os
from datetime import datetime
from typing import List, Dict
from .data_loader import LotteryData, LOTTERY_CONFIGS


class DataExporter:
    """数据导出器"""

    EXPORT_DIR = "exports"

    def __init__(self):
        os.makedirs(self.EXPORT_DIR, exist_ok=True)

    def _get_prefix(self, lottery_type: str = "dlt") -> str:
        return LOTTERY_CONFIGS.get(lottery_type, LOTTERY_CONFIGS["dlt"])["name"]

    def _get_headers(self, lottery_type: str = "dlt") -> list:
        config = LOTTERY_CONFIGS.get(lottery_type, LOTTERY_CONFIGS["dlt"])
        fn = config["front_name"]
        bn = config["back_name"]
        fc = config["front_count"]
        bc = config["back_count"]
        front_headers = [f"{fn}{i+1}" for i in range(fc)]
        back_headers = [f"{bn}{i+1}" for i in range(bc)]
        return ["期号", "开奖日期"] + front_headers + back_headers + [f"{fn}和值", f"{bn}和值"]

    def export_csv(self, data: List[LotteryData], filename: str = None, lottery_type: str = "dlt") -> str:
        """导出为CSV格式"""
        prefix = self._get_prefix(lottery_type)
        if filename is None:
            filename = f"{prefix}数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

        filepath = os.path.join(self.EXPORT_DIR, filename)
        with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(self._get_headers(lottery_type))
            for d in data:
                writer.writerow([
                    d.issue, d.date,
                    *d.front_nums, *d.back_nums,
                    sum(d.front_nums), sum(d.back_nums)
                ])
        return filepath

    def export_json(self, data: List[LotteryData], filename: str = None, lottery_type: str = "dlt") -> str:
        """导出为JSON格式"""
        prefix = self._get_prefix(lottery_type)
        if filename is None:
            filename = f"{prefix}数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

        filepath = os.path.join(self.EXPORT_DIR, filename)
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump([d.to_dict() for d in data], f, ensure_ascii=False, indent=2)
        return filepath

    def export_analysis_report(self, analysis_data: Dict, filename: str = None, lottery_type: str = "dlt") -> str:
        """导出分析报告为JSON"""
        prefix = self._get_prefix(lottery_type)
        if filename is None:
            filename = f"{prefix}分析报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

        filepath = os.path.join(self.EXPORT_DIR, filename)
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(analysis_data, f, ensure_ascii=False, indent=2)
        return filepath

    def export_excel(self, data: List[LotteryData], filename: str = None, lottery_type: str = "dlt") -> str:
        """导出为Excel格式（需要openpyxl）"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            prefix = self._get_prefix(lottery_type)
            csv_name = filename.replace(".xlsx", ".csv") if filename else None
            return self.export_csv(data, csv_name, lottery_type)

        prefix = self._get_prefix(lottery_type)
        if filename is None:
            filename = f"{prefix}数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        filepath = os.path.join(self.EXPORT_DIR, filename)
        wb = Workbook()
        ws = wb.active
        ws.title = "开奖数据"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center")

        headers = self._get_headers(lottery_type)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for row, d in enumerate(data, 2):
            values = [d.issue, d.date, *d.front_nums, *d.back_nums,
                      sum(d.front_nums), sum(d.back_nums)]
            for col, val in enumerate(values, 1):
                ws.cell(row=row, column=col, value=val)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 12

        wb.save(filepath)
        return filepath